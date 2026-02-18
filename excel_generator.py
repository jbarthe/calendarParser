import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_gantt(df_leaves):
    """
    Generates an Excel file with a Gantt chart layout.
    df_leaves: DataFrame [Name, Team, Start, End, Label]
    Returns: BytesIO object containing the Excel file.
    """
    if df_leaves.empty:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning Congés"

    # 1. Determine Date Range
    # Access as .dt accessor if it's a series, but here it's a scalar extraction?
    # Actually df_leaves['Start'] is a Series.
    min_date = df_leaves['Start'].min().normalize()
    max_date = df_leaves['End'].max().normalize()
    
    # Add buffer? Maybe not needed for Excel, just exact range or full months.
    # Let's go fully from 1st of min_month to end of max_month for cleaner look.
    start_date = min_date.replace(day=1)
    end_date = (max_date + pd.DateOffset(months=1)).replace(day=1) - pd.DateOffset(days=1)
    
    date_range = pd.date_range(start=start_date, end=end_date)
    total_days = len(date_range)

    # 2. Setup Styles
    # Colors
    from colors import assign_colors_by_team, COLOR_PALETTES
    person_color_map = assign_colors_by_team(df_leaves)
    
    # Clean up hex for OpenPyXL (remove #)
    person_color_map_clean = {k: v.lstrip('#') for k, v in person_color_map.items()}
    
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    month_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    weekend_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    border_side = Side(style='thin', color='000000')
    border_all = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    team_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    team_font = Font(bold=True)

    # 3. Draw Header (Timeline)
    # Row 1: Months
    # Row 2: Days
    
    ws.cell(row=2, column=1, value="Nom / Date").font = Font(bold=True)
    ws.column_dimensions['A'].width = 25
    
    col_idx = 2
    current_month = None
    month_start_col = 2
    
    for day in date_range:
        # Day Header
        cell = ws.cell(row=2, column=col_idx, value=day.day)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_all
        
        # Weekend styling
        if day.weekday() >= 5: # 5=Sat, 6=Sun
            cell.fill = weekend_fill
            
        # Set column width small (User requested ~9px, down from ~18px)
        ws.column_dimensions[get_column_letter(col_idx)].width = 1.5
        
        # Month Header Logic
        if current_month != day.month:
            if current_month is not None:
                # Merge previous month
                ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=col_idx-1)
                m_cell = ws.cell(row=1, column=month_start_col, value=date_range[month_start_col-2].strftime("%B %Y"))
                m_cell.alignment = Alignment(horizontal='center')
                m_cell.fill = header_fill
                m_cell.font = header_font
                
            current_month = day.month
            month_start_col = col_idx
            
        col_idx += 1
        
    # Merge last month
    ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=col_idx-1)
    m_cell = ws.cell(row=1, column=month_start_col, value=date_range[month_start_col-2].strftime("%B %Y"))
    m_cell.alignment = Alignment(horizontal='center')
    m_cell.fill = header_fill
    m_cell.font = header_font

    # 4. Draw Rows (Teams & People)
    # Prepare data structure similar to visualizer
    teams = df_leaves['Team'].unique()
    row_idx = 3
    
    # Map person to color index
    people_unique = df_leaves['Name'].unique()
    # person_color_map calculated above

    date_to_col = {d: i+2 for i, d in enumerate(date_range)}

    for team in teams:
        # Team Header
        t_cell = ws.cell(row=row_idx, column=1, value=team)
        t_cell.fill = team_fill
        t_cell.font = team_font
        t_cell.border = border_all
        
        # Fill the whole row for the team line? check if desired. 
        # Maybe just the name for now, or merge across?
        # Let's merge across all dates for the team separator line
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_days+1)
        t_cell.alignment = Alignment(horizontal='left', indent=1)
        
        row_idx += 1
        
        team_data = df_leaves[df_leaves['Team'] == team]
        people = team_data['Name'].unique()
        
        for person in people:
            # Name
            n_cell = ws.cell(row=row_idx, column=1, value=person)
            n_cell.border = border_all
            n_cell.font = Font(bold=True)
            
            # Draw Leaves
            leaves = team_data[team_data['Name'] == person]
            
            for _, leave in leaves.iterrows():
                # Find column start/end
                # We need to handle bounds if leave exceeds date_range (unlikely with our logic but safe to check)
                # Intersect leave range with chart range
                l_start = max(leave['Start'].normalize(), start_date)
                l_end = min(leave['End'].normalize(), end_date)
                
                if l_start > l_end:
                    continue
                
                # Get column indices
                try:
                    c_start = date_to_col[l_start]
                    # We need every individual day? No, just start and end for merge
                    # But date_to_col only has keys for specific timestamps (00:00:00).
                    # Ensure normalize
                    c_end = date_to_col[l_end]
                except KeyError:
                    continue

                # Style
                fill_color = person_color_map_clean[person]
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Merge
                if c_end > c_start:
                    ws.merge_cells(start_row=row_idx, start_column=c_start, end_row=row_idx, end_column=c_end)
                
                # Cell Content
                l_cell = ws.cell(row=row_idx, column=c_start, value=leave['Label'])
                l_cell.fill = fill
                l_cell.alignment = Alignment(horizontal='center')
                # Add borders to the merged range? Openpyxl styling on merged cells needs careful handling
                # Set border for all cells in range
                for c in range(c_start, c_end + 1):
                    ws.cell(row=row_idx, column=c).border = border_all
                    ws.cell(row=row_idx, column=c).fill = fill
            
            # Fill weekends for this row (if no leave)
            # This is expensive visually, maybe minimal is better.
            # Let's stick to leave coloring only for clarity.
            
            row_idx += 1

    # Freeze panes
    ws.freeze_panes = "B3"
    
    return wb

if __name__ == "__main__":
    # Test stub
    pass
