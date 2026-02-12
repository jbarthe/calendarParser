from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Saisie Congés"

    # Define Styles
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid") # Indigo
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    team_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid") # Lavender
    team_font = Font(color="000000", bold=True, size=11)
    
    border_style = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # 1. Main Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "📄 FORMULAIRE DE SAISIE DES CONGÉS"
    cell.font = Font(size=16, bold=True, color="4B0082")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 2. Instructions
    ws.merge_cells("A2:F2")
    ws["A2"].value = "Instructions : Remplissez les périodes sous la forme 'Du JJ/MM/AA au JJ/MM/AA'. Pour les jours isolés, utilisez '(+N JS : JJ/MM/AA)'."
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25

    # 3. Column Headers
    headers = ["Nom / Équipe", "Période 1", "Période 2", "Période 3", "Période 4", "Commentaires"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 4. Example Data
    data = [
        ("ADMINISTRATION", "", "", "", "", ""),
        ("Dupont Jean", "Du 01/01/25 au 05/01/25", "Du 10/06/25 au 20/06/25", "", "", ""),
        ("Durand Marie", "(+2 JS : 24 et 25/02/26)", "Du 01/08/25 au 15/08/25", "", "", ""),
        ("SERVICE TECHNIQUE", "", "", "", "", ""),
        ("Martin Paul", "Du 14/07/25 au 25/07/25", "(+1 JS : 02/05/26)", "", "", ""),
        ("Bernard Sophie", "Du 20/12/25 au 02/01/26", "", "", "", "")
    ]

    start_row = 4
    for i, row_data in enumerate(data):
        current_row = start_row + i
        
        # Check if it's a team header (first col has text, others empty)
        is_team = row_data[0].isupper() and not any(row_data[1:])
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.border = thin_border
            
            if is_team:
                cell.fill = team_fill
                cell.font = team_font
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="left", indent=1)
            else:
                cell.alignment = Alignment(horizontal="left")

    # 5. Adjust Column Widths
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F"]:
         ws.column_dimensions[col].width = 25

    # Save
    wb.save("template.xlsx")
    print("template.xlsx created successfully.")

if __name__ == "__main__":
    create_styled_template()
